#!/usr/bin/env python
# -*- coding:utf-8 -*-

import os, os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from shutil import copyfile

'''
1.读取excel文件到数组中
2.对比老的文件夹数组中文件名，若相同将文件将其存入新的文件夹
'''
book = load_workbook(filename=r"excel.xlsx")
sheet = book.get_sheet_by_name("Sheet1")
total_row = sheet.max_row #统计excel的行数
data = []
row_num = 1
while row_num <= total_row:
    data.append(sheet.cell(row=row_num, column=1).value)
    row_num += 1


def file_to_Other_Dir(old_path, new_path):

    for root,dirs,files in os.walk(old_path):
        for file in files:
            if file in data:
                # 复制文件到新的文件夹
                # new_path = new_path + '/'
                # copyfile(old_path+'/'+file, new_path+file)
                copyfile(file, new_path+file)
                # 剪切文件到文件夹
                oldpath = os.path.join(old_path, file)
                newpath = os.path.join(new_path, file)
                print(oldpath)
                print(newpath)
                os.rename(oldpath, newpath)  #相当于剪切


if __name__ == "__main__":
    old_path = './老的文件夹文件'
    new_path = './same_with_excel'
    file_to_Other_Dir(old_path, new_path)
